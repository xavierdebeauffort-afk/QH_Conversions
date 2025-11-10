# Updated Energy Consumption Analyzer with All Improvements

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles
from PIL import Image
from datetime import datetime
import io
import numpy as np
import streamlit as st
import tempfile
import os
import hashlib
import time
from typing import Optional, Tuple, Dict, Any, List

# ============================================
# CONFIGURATION
# ============================================

class Config:
    """Application configuration"""
    
    # File handling
    MAX_FILE_SIZE_MB = 50
    SUPPORTED_FORMATS = ['xlsx', 'csv']
    
    # Data processing
    EXPECTED_INTERVAL_MINUTES = 15
    CHUNK_SIZE = 10000
    
    # Visualization
    LUMINUS_GREEN = '#00B612'
    SEASON_COLORS = {
        'Spring': '#DA32A4',
        'Summer': '#007BFF', 
        'Fall': '#FF7F11',
        'Winter': '#D72638'
    }
    
    # Excel output
    PLOT_DPI = 100
    IMAGE_ANCHOR_COL = 'E1'

# ============================================
# CUSTOM CSS FOR BETTER UI
# ============================================

st.markdown("""
<style>
    /* Main containers */
    .block-container {
        padding-top: 2rem;
        max-width: 1200px;
    }
    
    /* Info boxes */
    .stAlert {
        border-radius: 0.5rem;
    }
    
    /* Metrics */
    [data-testid="stMetricValue"] {
        font-size: 1.5rem;
    }
    
    /* File uploader - Dark theme */
    [data-testid="stFileUploader"] {
        border: 2px dashed #4a90e2;
        border-radius: 0.5rem;
        padding: 2rem;
        background: #1e1e1e;
    }
    
    [data-testid="stFileUploader"] section {
        background: #2d2d2d;
        border-radius: 0.5rem;
        padding: 2rem;
    }
    
    [data-testid="stFileUploader"] section > div {
        color: #ffffff;
    }
    
    /* Upload text styling */
    [data-testid="stFileUploader"] label {
        color: #ffffff !important;
    }
    
    /* Browse files button */
    [data-testid="stFileUploader"] button {
        background: #2d2d2d;
        color: #ffffff;
        border: 1px solid #4a90e2;
        border-radius: 0.5rem;
        padding: 0.5rem 1rem;
    }
    
    [data-testid="stFileUploader"] button:hover {
        background: #3d3d3d;
        border-color: #5fa3f5;
    }
    
    /* Buttons */
    .stButton > button {
        border-radius: 0.5rem;
        font-weight: 600;
    }
    
    /* Header styling */
    .main-title {
        text-align: center;
        padding: 2rem 0 1rem 0;
    }
    
    .main-title h1 {
        color: #1f77b4;
        margin-bottom: 0.5rem;
    }
    
    .subtitle {
        text-align: center;
        font-size: 1.2rem;
        color: #888;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# UTILITY FUNCTIONS
# ============================================

def get_file_hash(file_content: bytes) -> str:
    """Generate hash for file content to use as cache key"""
    return hashlib.md5(file_content).hexdigest()

def initialize_session_state():
    """Initialize session state variables"""
    if 'results' not in st.session_state:
        st.session_state.results = None
    if 'file_processed' not in st.session_state:
        st.session_state.file_processed = False
    if 'last_file_hash' not in st.session_state:
        st.session_state.last_file_hash = None

# ============================================
# DATA INGESTION FUNCTIONS
# ============================================

def validate_uploaded_file(uploaded_file) -> bool:
    """Validate file before processing"""
    try:
        file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
        if file_size_mb > Config.MAX_FILE_SIZE_MB:
            st.error(f"❌ File too large: {file_size_mb:.1f}MB. Maximum: {Config.MAX_FILE_SIZE_MB}MB")
            return False
        
        file_ext = uploaded_file.name.split('.')[-1].lower()
        if file_ext not in Config.SUPPORTED_FORMATS:
            st.error(f"❌ Unsupported format: .{file_ext}. Supported: {', '.join(Config.SUPPORTED_FORMATS)}")
            return False
        
        if file_size_mb < 0.001:
            st.error("❌ File appears to be empty")
            return False
        
        return True
    
    except Exception as e:
        st.error(f"❌ File validation error: {str(e)}")
        return False

def clean_numeric_column(series: pd.Series) -> pd.Series:
    """Convert European or US number formats to float"""
    if series.dtype == 'object':
        cleaned = series.astype(str).str.replace(',', '.').str.replace(' ', '').str.strip()
        return pd.to_numeric(cleaned, errors='coerce')
    return series

def parse_datetime_robust(date_str: str, time_str: str) -> pd.Timestamp:
    """Try multiple datetime formats (US and EU)"""
    formats_to_try = [
        '%d.%m.%Y %H:%M:%S',
        '%m/%d/%Y %I:%M:%S %p',
        '%m/%d/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%d.%m.%Y %H:%M',
        '%m/%d/%Y %H:%M',
    ]
    
    combined = f"{str(date_str).strip()} {str(time_str).strip()}"
    
    for fmt in formats_to_try:
        try:
            return pd.to_datetime(combined, format=fmt)
        except:
            continue
    
    try:
        return pd.to_datetime(combined, infer_datetime_format=True)
    except:
        return pd.NaT

def detect_delimiter(content: str) -> str:
    """Detect CSV delimiter"""
    first_line = content.split('\n')[0]
    delimiters = [';', ',', '\t']
    
    for delim in delimiters:
        if delim in first_line:
            return delim
    
    return ','

def read_consumption_file(uploaded_file) -> pd.DataFrame:
    """Read Excel or CSV file with detailed error handling"""
    file_ext = uploaded_file.name.split('.')[-1].lower()
    
    try:
        if file_ext == 'xlsx':
            try:
                df = pd.read_excel(uploaded_file, sheet_name='Data_RAW', 
                                 header=None, engine='openpyxl')
            except ValueError as e:
                if "Worksheet" in str(e):
                    uploaded_file.seek(0)
                    df = pd.read_excel(uploaded_file, header=None, engine='openpyxl')
                    st.warning("⚠️ 'Data_RAW' sheet not found. Using first sheet.")
                else:
                    raise
        
        elif file_ext == 'csv':
            content = uploaded_file.getvalue().decode('utf-8-sig')
            delimiter = detect_delimiter(content)
            
            uploaded_file.seek(0)
            
            try:
                df = pd.read_csv(uploaded_file, delimiter=delimiter, 
                               header=None, decimal='.', encoding='utf-8-sig')
            except Exception:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, delimiter=delimiter, 
                               header=None, decimal=',', encoding='utf-8-sig')
        
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        return df
    
    except Exception as e:
        st.error(f"❌ File reading error: {str(e)}")
        raise

def validate_timeseries(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Check for data quality issues in time series"""
    issues = []
    
    duplicates = df[df['DATETIME'].duplicated()]
    if len(duplicates) > 0:
        issues.append({
            'type': 'DUPLICATE_TIMESTAMPS',
            'count': len(duplicates),
            'message': f"Found {len(duplicates)} duplicate timestamps",
            'action': 'Will keep first occurrence, remove duplicates'
        })
    
    df_sorted = df.sort_values('DATETIME')
    time_diff = df_sorted['DATETIME'].diff()
    expected_interval = pd.Timedelta(minutes=Config.EXPECTED_INTERVAL_MINUTES)
    
    gaps = time_diff[time_diff > expected_interval * 1.5]
    if len(gaps) > 0:
        total_missing = gaps.sum() / expected_interval
        issues.append({
            'type': 'MISSING_INTERVALS',
            'count': len(gaps),
            'message': f"Found {len(gaps)} gaps in data (approx {int(total_missing)} missing intervals)",
            'action': 'Gaps will be handled during hourly resampling'
        })
    
    non_standard = time_diff[(time_diff != expected_interval) & 
                            (time_diff.notna()) & 
                            (time_diff < expected_interval * 1.5)]
    if len(non_standard) > 5:
        issues.append({
            'type': 'IRREGULAR_INTERVALS',
            'count': len(non_standard),
            'message': f"Found {len(non_standard)} irregular time intervals",
            'action': 'Will be handled during hourly resampling'
        })
    
    if (df[2] < 0).any():
        neg_count = (df[2] < 0).sum()
        issues.append({
            'type': 'NEGATIVE_VALUES',
            'count': neg_count,
            'message': f"Found {neg_count} negative consumption values",
            'action': 'Negative values will be set to 0'
        })
    
    null_dates = df['DATETIME'].isna().sum()
    null_consumption = df[2].isna().sum()
    if null_dates > 0 or null_consumption > 0:
        issues.append({
            'type': 'NULL_VALUES',
            'count': null_dates + null_consumption,
            'message': f"Found {null_dates} null dates and {null_consumption} null consumption values",
            'action': 'Rows with null values will be removed'
        })
    
    return issues

def show_data_quality_report(issues: List[Dict[str, Any]]) -> bool:
    """Display data quality issues in Streamlit"""
    if not issues:
        st.success("✅ No data quality issues detected! Data is clean and ready for analysis.")
        return True
    
    st.warning(f"⚠️ Data Quality Issues Detected ({len(issues)} issues)")
    
    for i, issue in enumerate(issues, 1):
        with st.expander(f"Issue {i}: {issue['type']}", expanded=True):
            st.write(f"**{issue['message']}**")
            st.info(f"Action: {issue['action']}")
    
    st.write("---")
    return True

def clean_timeseries(df: pd.DataFrame, issues: List[Dict[str, Any]]) -> pd.DataFrame:
    """Apply cleaning based on identified issues"""
    df_clean = df.copy()
    
    for issue in issues:
        if issue['type'] == 'DUPLICATE_TIMESTAMPS':
            df_clean = df_clean.drop_duplicates(subset='DATETIME', keep='first')
        
        elif issue['type'] == 'NEGATIVE_VALUES':
            df_clean[2] = df_clean[2].clip(lower=0)
        
        elif issue['type'] == 'NULL_VALUES':
            df_clean = df_clean.dropna(subset=['DATETIME', 2])
    
    return df_clean.sort_values('DATETIME').reset_index(drop=True)

def filter_by_date_range(df: pd.DataFrame, filter_mode: str, 
                        selected_year: int) -> Tuple[Optional[pd.DataFrame], str]:
    """Filter dataframe based on pre-selected date range"""
    min_date = df['DATETIME'].min()
    max_date = df['DATETIME'].max()
    available_years = sorted(df['DATETIME'].dt.year.unique())
    
    if filter_mode == 'all':
        return df, f"All data ({min_date.date()} to {max_date.date()})"
    
    elif filter_mode == 'recent':
        most_recent_year = available_years[-1]
        filtered = df[df['DATETIME'].dt.year == most_recent_year]
        
        if len(filtered) == 0:
            st.warning(f"⚠️ No complete data for {most_recent_year}")
            return None, f"No data for {most_recent_year}"
        
        return filtered, f"Year {most_recent_year} ({len(filtered):,} records)"
    
    elif filter_mode == 'custom':
        if selected_year not in available_years:
            st.error(f"❌ Year {selected_year} not available.")
            st.info(f"Available years: {', '.join(map(str, available_years))}")
            return None, f"Invalid year selection"
        
        filtered = df[df['DATETIME'].dt.year == selected_year]
        return filtered, f"Year {selected_year} ({len(filtered):,} records)"
    
    return df, "All data"

def calculate_load_duration_curve(hourly_df: pd.DataFrame) -> pd.DataFrame:
    """Generate load duration curve data"""
    consumption = hourly_df['Consumption [kWh]'].values
    sorted_consumption = np.sort(consumption)[::-1]
    hours = np.arange(1, len(sorted_consumption) + 1)
    percentiles = (hours / len(hours)) * 100
    
    return pd.DataFrame({
        'Hours': hours,
        'Consumption [kWh]': sorted_consumption,
        'Percentile': percentiles
    })

def get_season(month):
    """Get season from month number"""
    if month in [12, 1, 2]:
        return 'Winter'
    elif month in [3, 4, 5]:
        return 'Spring'
    elif month in [6, 7, 8]:
        return 'Summer'
    else:
        return 'Fall'

# ============================================
# VISUALIZATION FUNCTIONS
# ============================================

def create_heatmap(dataframe: pd.DataFrame, title: str) -> plt.Figure:
    """Create a weekly heatmap for hourly consumption"""
    df_heat = dataframe.copy()
    df_heat['Hour'] = df_heat['Timestamp'].dt.hour
    df_heat['Weekday'] = df_heat['Timestamp'].dt.day_name()
    pivot = df_heat.groupby(['Weekday', 'Hour'])['Consumption [kWh]'].mean().unstack()
    ordered_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    pivot = pivot.reindex(ordered_days)
    
    fig = plt.figure(figsize=(14, 6))
    sns.heatmap(pivot, cmap='RdYlGn_r', annot=True, fmt=".0f", 
                linewidths=0.5, cbar_kws={'label': 'Avg kWh'})
    plt.title(title, fontsize=14, fontweight='bold')
    plt.xlabel('Hour of Day')
    plt.ylabel('Day of Week')
    plt.tight_layout()
    return fig

def create_monthly_stats_charts(hourly_df: pd.DataFrame) -> Tuple[plt.Figure, plt.Figure, pd.DataFrame, pd.DataFrame]:
    """Create monthly min/max statistics visualizations"""
    df = hourly_df.copy()
    df['Month'] = df['Timestamp'].dt.strftime('%b %Y')
    df['MonthStart'] = df['Timestamp'].dt.to_period('M').dt.to_timestamp()
    
    monthly_stats = df.groupby(['Month', 'MonthStart'])['Consumption [kWh]'].agg(['min', 'max', 'mean']).reset_index()
    monthly_stats.columns = ['Month', 'MonthStart', 'Min', 'Max', 'Average']
    monthly_stats = monthly_stats.sort_values('MonthStart').reset_index(drop=True)
    
    # Line chart
    fig1 = plt.figure(figsize=(14, 5))
    plt.plot(monthly_stats['Month'], monthly_stats['Min'], marker='o', label='Min', color='#4CAF50', linewidth=2)
    plt.plot(monthly_stats['Month'], monthly_stats['Max'], marker='o', label='Max', color='#B22222', linewidth=2)
    plt.plot(monthly_stats['Month'], monthly_stats['Average'], marker='o', label='Average', color='gray', linewidth=2)
    plt.title('Monthly Min, Max, and Average Consumption', fontsize=14, fontweight='bold')
    plt.xticks(rotation=45, ha='right')
    plt.ylabel('Consumption [kWh]')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    # Heatmap
    heatmap_data = pd.DataFrame({
        'Max': monthly_stats['Max'].values,
        'Average': monthly_stats['Average'].values,
        'Min': monthly_stats['Min'].values
    }, index=monthly_stats['Month']).T
    
    monthly_stats['Quarter'] = monthly_stats['MonthStart'].dt.to_period('Q')
    quarter_groups = monthly_stats.groupby('Quarter')
    highlight_cells = {'Max': [], 'Min': []}
    quarter_boundaries = []
    
    for quarter, group in quarter_groups:
        max_month = group.loc[group['Max'].idxmax(), 'Month']
        min_month = group.loc[group['Min'].idxmin(), 'Month']
        highlight_cells['Max'].append(max_month)
        highlight_cells['Min'].append(min_month)
        last_month_index = monthly_stats[monthly_stats['Month'] == group['Month'].iloc[-1]].index[0]
        quarter_boundaries.append(last_month_index + 1)
    
    fig2 = plt.figure(figsize=(14, 3))
    ax = sns.heatmap(heatmap_data, annot=True, fmt=".0f", linewidths=0.5, linecolor='gray',
                    cbar=False, xticklabels=True, yticklabels=True, cmap='Greys',
                    mask=heatmap_data.isnull(), alpha=0)
    
    for y_index, row in enumerate(heatmap_data.index):
        for x_index, month in enumerate(heatmap_data.columns):
            if row == 'Max' and month in highlight_cells['Max']:
                ax.add_patch(plt.Rectangle((x_index, y_index), 1, 1, fill=True, color='#B22222', linewidth=0))
            elif row == 'Min' and month in highlight_cells['Min']:
                ax.add_patch(plt.Rectangle((x_index, y_index), 1, 1, fill=True, color='#4CAF50', linewidth=0))
            ax.text(x_index + 0.5, y_index + 0.5, f"{heatmap_data.loc[row, month]:.0f}",
                    ha='center', va='center', color='black', fontweight='bold')
    
    for y in range(1, heatmap_data.shape[0]):
        ax.axhline(y=y, color='gray', linewidth=0.5, alpha=0.3)
    for x in quarter_boundaries:
        if x < heatmap_data.shape[1]:
            ax.axvline(x=x, color='black', linewidth=1.5)
    
    ax.add_patch(plt.Rectangle((0, 0), heatmap_data.shape[1], heatmap_data.shape[0],
                               fill=False, edgecolor='black', lw=1.5))
    plt.title('Monthly Consumption Statistics – Highlighted Heatmap', fontsize=14, fontweight='bold')
    plt.xlabel('')
    plt.ylabel('')
    plt.xticks(rotation=45, ha='right')
    plt.yticks(rotation=0)
    plt.tight_layout()
    
    # Quarterly stats
    quarterly_stats = monthly_stats.groupby('Quarter').agg({'Min': 'min', 'Max': 'max'}).reset_index()
    quarterly_stats['Time'] = quarterly_stats['Quarter'].astype(str)
    quarterly_stats = quarterly_stats[['Time', 'Min', 'Max']]
    
    return fig1, fig2, monthly_stats[['Month', 'Min', 'Max', 'Average']], quarterly_stats

def generate_all_plots(hourly_df: pd.DataFrame, monthly_df: pd.DataFrame, 
                      raw_df: pd.DataFrame) -> Dict[str, Any]:
    """Generate all visualizations and return as dict of figures"""
    plots = {}
    
    df = hourly_df.copy()
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    
    # 1. Histogram
    fig1 = plt.figure(figsize=(14, 6))
    sns.histplot(df['Consumption [kWh]'], bins=50, kde=False, color=Config.LUMINUS_GREEN)
    plt.title('Histogram of Hourly Consumption', fontsize=14, fontweight='bold')
    plt.xlabel('Consumption [kWh]')
    plt.ylabel('Frequency')
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['histogram'] = fig1
    
    # 2. Time Series
    fig2 = plt.figure(figsize=(14, 6))
    plt.plot(df['Timestamp'], df['Consumption [kWh]'], color=Config.LUMINUS_GREEN, linewidth=0.8)
    plt.title('Hourly Consumption Over Time', fontsize=14, fontweight='bold')
    plt.xlabel('Date')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b'))
    plt.tight_layout()
    plots['timeseries'] = fig2
    
    # 3. Daily Profile
    df['Hour'] = df['Timestamp'].dt.hour
    avg_hourly = df.groupby('Hour')['Consumption [kWh]'].mean()
    fig3 = plt.figure(figsize=(14, 6))
    avg_hourly.plot(kind='bar', color=Config.LUMINUS_GREEN, width=0.9)
    plt.title('Average Daily Consumption Profile', fontsize=14, fontweight='bold')
    plt.xlabel('Hour of Day')
    plt.ylabel('Average Consumption [kWh]')
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['daily_profile'] = fig3
    
    # 4. Heatmaps (Yearly, January, July)
    plots['heatmap_yearly'] = create_heatmap(df, 'Average Hourly Consumption (Yearly)')
    
    df_jan = df[df['Timestamp'].dt.month == 1]
    if len(df_jan) > 0:
        plots['heatmap_january'] = create_heatmap(df_jan, 'Average Hourly Consumption - January')
    
    df_jul = df[df['Timestamp'].dt.month == 7]
    if len(df_jul) > 0:
        plots['heatmap_july'] = create_heatmap(df_jul, 'Average Hourly Consumption - July')
    
    # 5. Seasonal Patterns
    df['Season'] = df['Timestamp'].dt.month.map(get_season)
    seasonal_hourly = df.groupby(['Hour', 'Season'])['Consumption [kWh]'].mean().unstack()
    
    fig4 = plt.figure(figsize=(14, 6))
    for season in ['Winter', 'Spring', 'Summer', 'Fall']:
        if season in seasonal_hourly.columns:
            plt.plot(seasonal_hourly.index, seasonal_hourly[season], 
                    label=season, color=Config.SEASON_COLORS[season], linewidth=2)
    plt.title('Average Hourly Consumption by Season', fontsize=14, fontweight='bold')
    plt.xlabel('Hour of Day')
    plt.ylabel('Average Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plots['seasonal'] = fig4
    
    # 6. Monthly Consumption Bar Chart
    df_monthly = monthly_df.copy()
    fig5 = plt.figure(figsize=(14, 6))
    sns.barplot(x='Month', y='Consumption [kWh]', data=df_monthly, color=Config.LUMINUS_GREEN)
    plt.title('Monthly Energy Consumption', fontsize=14, fontweight='bold')
    plt.xlabel('Month')
    plt.ylabel('Consumption [kWh]')
    plt.xticks(rotation=45)
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['monthly_bar'] = fig5
    
    # 7. Capacity Tariff - Top 10 Peaks
    df_raw_analysis = raw_df.copy()
    df_raw_analysis = df_raw_analysis.rename(columns={'Consumption': 'Consumption [kWh]'})
    
    top10 = df_raw_analysis.nlargest(10, 'Consumption [kWh]').sort_values('Timestamp')
    
    fig6 = plt.figure(figsize=(14, 6))
    plt.plot(df_raw_analysis['Timestamp'], df_raw_analysis['Consumption [kWh]'], 
            color=Config.LUMINUS_GREEN, linewidth=0.8, label='Consumption', alpha=0.7)
    plt.scatter(top10['Timestamp'], top10['Consumption [kWh]'], 
               color='red', s=100, label='Top 10 Peaks', zorder=5)
    plt.title('Time Series with Top 10 Peaks', fontsize=14, fontweight='bold')
    plt.xlabel('Date')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b'))
    plt.legend()
    plt.tight_layout()
    plots['top10_peaks'] = fig6
    
    # 8. Top 100 by Season
    top100 = df_raw_analysis.nlargest(100, 'Consumption [kWh]').copy()
    top100['Season'] = top100['Timestamp'].dt.month.map(get_season)
    season_counts = top100['Season'].value_counts().reindex(['Spring', 'Summer', 'Fall', 'Winter'])
    
    fig7 = plt.figure(figsize=(10, 6))
    season_counts.plot(kind='bar', color=[Config.SEASON_COLORS[s] for s in season_counts.index])
    plt.title('Top 100 Consumption Values by Season', fontsize=14, fontweight='bold')
    plt.xlabel('Season')
    plt.ylabel('Count')
    plt.xticks(rotation=0)
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['top100_seasons'] = fig7
    
    # 9. Monthly Min/Max Statistics
    fig_monthly_line, fig_monthly_heatmap, monthly_table, quarterly_table = create_monthly_stats_charts(hourly_df)
    plots['monthly_minmax_line'] = fig_monthly_line
    plots['monthly_minmax_heatmap'] = fig_monthly_heatmap
    plots['monthly_stats_table'] = monthly_table
    plots['quarterly_stats_table'] = quarterly_table
    
    # 10. Load Duration Curve
    ldc_data = calculate_load_duration_curve(hourly_df)
    fig8 = plt.figure(figsize=(14, 6))
    plt.plot(ldc_data['Percentile'], ldc_data['Consumption [kWh]'], 
            color=Config.LUMINUS_GREEN, linewidth=2)
    plt.axhline(y=hourly_df['Consumption [kWh]'].mean(), color='red', 
               linestyle='--', label=f"Average ({hourly_df['Consumption [kWh]'].mean():.1f} kWh)")
    plt.axhline(y=hourly_df['Consumption [kWh]'].median(), color='orange', 
               linestyle='--', label=f"Median ({hourly_df['Consumption [kWh]'].median():.1f} kWh)")
    plt.title('Load Duration Curve', fontsize=14, fontweight='bold')
    plt.xlabel('Percentage of Time (%)')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plots['load_duration'] = fig8
    
    # Store additional data
    plots['seasonal_hourly_data'] = seasonal_hourly
    plots['top10_data'] = top10
    plots['top100_data'] = top100
    plots['ldc_data'] = ldc_data
    
    return plots

def create_excel_file(hourly_df: pd.DataFrame, monthly_df: pd.DataFrame, 
                     raw_df: pd.DataFrame, plots: Dict[str, Any], 
                     filename: str) -> io.BytesIO:
    """Create Excel file in memory with all data and visualizations"""
    
    hourly_df_no_feb29 = hourly_df[~((hourly_df['Timestamp'].dt.month == 2) & 
                                     (hourly_df['Timestamp'].dt.day == 29))].reset_index(drop=True)
    
    raw_df_for_save = raw_df[['Date', 'Time', 'Consumption']].copy()
    
    with tempfile.TemporaryDirectory() as tmpdir:
        plot_files = {}
        
        for name, fig in plots.items():
            if name.endswith('_data') or name.endswith('_table'):
                continue
            filepath = os.path.join(tmpdir, f'{name}.png')
            fig.savefig(filepath, dpi=Config.PLOT_DPI, bbox_inches='tight')
            plot_files[name] = filepath
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            hourly_df.to_excel(writer, sheet_name='Hourly', index=False)
            monthly_df.to_excel(writer, sheet_name='Monthly', index=False)
            hourly_df_no_feb29.to_excel(writer, sheet_name='Sympheny', index=False)
            raw_df_for_save.to_excel(writer, sheet_name='Data_RAW', index=False, header=False)
        
        output.seek(0)
        wb = load_workbook(output)
        
        def add_image_to_sheet(sheet_name, image_path, anchor='A1'):
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]
            img = XLImage(image_path)
            img.anchor = anchor
            ws.add_image(img)
        
        # Add all visualizations
        add_image_to_sheet('Histogram', plot_files['histogram'])
        add_image_to_sheet('Time Series Yearly', plot_files['timeseries'])
        add_image_to_sheet('Daily Profile', plot_files['daily_profile'])
        
        # Heatmaps
        add_image_to_sheet('Hourly Heatmap', plot_files['heatmap_yearly'])
        if 'heatmap_january' in plot_files:
            add_image_to_sheet('Heatmap January', plot_files['heatmap_january'])
        if 'heatmap_july' in plot_files:
            add_image_to_sheet('Heatmap July', plot_files['heatmap_july'])
        
        # Seasonal patterns
        ws_seasonal = wb.create_sheet(title='Weekly patterns')
        seasonal_data = plots['seasonal_hourly_data'].reset_index()
        for r_idx, row in enumerate(dataframe_to_rows(seasonal_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_seasonal.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['seasonal'])
        img.anchor = 'F1'
        ws_seasonal.add_image(img)
        
        # Monthly chart
        ws_monthly = wb['Monthly']
        img = XLImage(plot_files['monthly_bar'])
        img.anchor = 'F1'
        ws_monthly.add_image(img)
        
        # Capacity tariff
        ws_top10 = wb.create_sheet(title='Quarter top 10')
        top10_data = plots['top10_data'][['Timestamp', 'Consumption [kWh]']]
        for r_idx, row in enumerate(dataframe_to_rows(top10_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_top10.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['top10_peaks'])
        img.anchor = 'E1'
        ws_top10.add_image(img)
        
        ws_top100 = wb.create_sheet(title='Quarter top 100')
        top100_data = plots['top100_data'][['Timestamp', 'Consumption [kWh]', 'Season']]
        for r_idx, row in enumerate(dataframe_to_rows(top100_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_top100.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['top100_seasons'])
        img.anchor = 'E1'
        ws_top100.add_image(img)
        
        # Monthly stats
        ws_monthly_minmax = wb.create_sheet("MinMax_Monthly")
        for r_idx, row in enumerate(dataframe_to_rows(plots['monthly_stats_table'], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_monthly_minmax.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['monthly_minmax_line'])
        img.anchor = "F1"
        ws_monthly_minmax.add_image(img)
        
        ws_quarter = wb.create_sheet("MinMax_Quarter")
        for r_idx, row in enumerate(dataframe_to_rows(plots['quarterly_stats_table'], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_quarter.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['monthly_minmax_heatmap'])
        img.anchor = "F1"
        ws_quarter.add_image(img)
        
        # Load duration curve
        ws_ldc = wb.create_sheet(title='Load Duration Curve')
        ldc_data = plots['ldc_data']
        for r_idx, row in enumerate(dataframe_to_rows(ldc_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_ldc.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['load_duration'])
        img.anchor = 'E1'
        ws_ldc.add_image(img)
        
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output

# ============================================
# MAIN PROCESSING FUNCTION
# ============================================

@st.cache_data(show_spinner=False)
def process_data_streamlit(file_hash: str, file_content: bytes, file_name: str, 
                          filter_mode: str, selected_year: int) -> Tuple[Optional[Dict[str, Any]], None, Optional[str]]:
    """Main processing function with progress tracking"""
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        status_text.text("📂 Loading data...")
        uploaded_file = io.BytesIO(file_content)
        uploaded_file.name = file_name
        df = read_consumption_file(uploaded_file)
        progress_bar.progress(20)
        
        status_text.text("🧹 Cleaning data...")
        df[2] = clean_numeric_column(df[2])
        df['DATETIME'] = df.apply(
            lambda row: parse_datetime_robust(row[0], row[1]), axis=1
        )
        progress_bar.progress(40)
        
        status_text.text("🔍 Validating data quality...")
        issues = validate_timeseries(df)
        df = clean_timeseries(df, issues)
        
        df, date_range_info = filter_by_date_range(df, filter_mode, selected_year)
        
        if df is None or len(df) == 0:
            progress_bar.empty()
            status_text.empty()
            return None, None, f"No data to process. {date_range_info}"
        
        progress_bar.progress(60)
        
        status_text.text("⚙️ Processing hourly data...")
        raw_df = df[[0, 1, 2, 'DATETIME']].copy()
        raw_df.columns = ['Date', 'Time', 'Consumption', 'Timestamp']
        
        df.set_index('DATETIME', inplace=True)
        hourly_df = df[2].resample('H').sum().reset_index()
        hourly_df[2] = hourly_df[2] / 4
        hourly_df.columns = ['Timestamp', 'Consumption [kWh]']
        
        monthly_df = hourly_df.copy()
        monthly_df['Month'] = monthly_df['Timestamp'].dt.to_period('M')
        monthly_df = monthly_df.groupby('Month')['Consumption [kWh]'].sum().reset_index()
        monthly_df['Month'] = monthly_df['Month'].astype(str)
        
        progress_bar.progress(70)
        
        status_text.text("📊 Creating visualizations...")
        plots = generate_all_plots(hourly_df, monthly_df, raw_df)
        progress_bar.progress(85)
        
        status_text.text("📄 Generating Excel report...")
        excel_file = create_excel_file(hourly_df, monthly_df, raw_df, plots, file_name)
        progress_bar.progress(100)
        
        status_text.text("✅ Complete!")
        time.sleep(0.5)
        progress_bar.empty()
        status_text.empty()
        
        return {
            'hourly_df': hourly_df,
            'monthly_df': monthly_df,
            'raw_df': raw_df,
            'plots': plots,
            'excel_file': excel_file,
            'date_range_info': date_range_info,
            'issues': issues
        }, None, None
    
    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        return None, None, f"Processing error: {str(e)}"

# ============================================
# UI DISPLAY FUNCTIONS
# ============================================

def display_summary_metrics(results: Dict[str, Any]):
    """Display key metrics at the top"""
    
    hourly_df = results['hourly_df']
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        total = hourly_df['Consumption [kWh]'].sum()
        st.metric("Total Consumption", f"{total:,.0f} kWh")
    
    with col2:
        avg = hourly_df['Consumption [kWh]'].mean()
        st.metric("Average (Hourly)", f"{avg:.1f} kWh")
    
    with col3:
        peak = hourly_df['Consumption [kWh]'].max()
        st.metric("Peak Demand", f"{peak:.1f} kWh")
    
    with col4:
        data_points = len(results['raw_df'])
        st.metric("Data Points", f"{data_points:,}")
    
    with col5:
        date_range = (hourly_df['Timestamp'].max() - 
                     hourly_df['Timestamp'].min()).days
        st.metric("Days Covered", f"{date_range}")

def display_visualizations(plots: Dict[str, Any]):
    """Display all plots in organized layout"""
    
    st.subheader("📊 Consumption Overview")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Histogram of Hourly Consumption**")
        st.pyplot(plots['histogram'])
    
    with col2:
        st.markdown("**Time Series - Yearly View**")
        st.pyplot(plots['timeseries'])
    
    st.divider()
    
    st.subheader("🔥 Hourly Consumption Heatmaps")
    
    st.markdown("**Yearly Average**")
    st.pyplot(plots['heatmap_yearly'])
    
    col3, col4 = st.columns(2)
    
    with col3:
        if 'heatmap_january' in plots:
            st.markdown("**January**")
            st.pyplot(plots['heatmap_january'])
    
    with col4:
        if 'heatmap_july' in plots:
            st.markdown("**July**")
            st.pyplot(plots['heatmap_july'])
    
    st.divider()
    
    st.subheader("📈 Consumption Patterns")
    col5, col6 = st.columns(2)
    
    with col5:
        st.markdown("**Average Daily Profile**")
        st.pyplot(plots['daily_profile'])
    
    with col6:
        st.markdown("**Seasonal Hourly Patterns**")
        st.pyplot(plots['seasonal'])
    
    st.divider()
    
    st.subheader("📅 Monthly Analysis")
    
    col7, col8 = st.columns(2)
    
    with col7:
        st.markdown("**Monthly Consumption**")
        st.pyplot(plots['monthly_bar'])
    
    with col8:
        st.markdown("**Monthly Min/Max Trends**")
        st.pyplot(plots['monthly_minmax_line'])
    
    st.markdown("**Quarterly Min/Max Heatmap**")
    st.pyplot(plots['monthly_minmax_heatmap'])
    
    st.divider()
    
    st.subheader("⚡ Peak Demand Analysis")
    col9, col10 = st.columns(2)
    
    with col9:
        st.markdown("**Top 10 Consumption Peaks**")
        st.pyplot(plots['top10_peaks'])
    
    with col10:
        st.markdown("**Top 100 Peaks by Season**")
        st.pyplot(plots['top100_seasons'])
    
    st.divider()
    
    st.subheader("📉 Load Duration Curve")
    st.pyplot(plots['load_duration'])

def display_data_tables(results: Dict[str, Any]):
    """Display data tables"""
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Hourly Data", 
        "Monthly Summary",
        "Monthly Stats",
        "Top 10 Peaks",
        "Load Duration"
    ])
    
    with tab1:
        st.dataframe(
            results['hourly_df'].style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True,
            height=400
        )
        st.caption(f"Total rows: {len(results['hourly_df'])}")
    
    with tab2:
        st.dataframe(
            results['monthly_df'].style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True
        )
        total_consumption = results['monthly_df']['Consumption [kWh]'].sum()
        st.metric("Total Consumption", f"{total_consumption:,.2f} kWh")
    
    with tab3:
        st.markdown("**Monthly Statistics**")
        st.dataframe(
            results['plots']['monthly_stats_table'].style.format({
                'Min': '{:.2f}',
                'Max': '{:.2f}',
                'Average': '{:.2f}'
            }),
            use_container_width=True
        )
        
        st.markdown("**Quarterly Statistics**")
        st.dataframe(
            results['plots']['quarterly_stats_table'].style.format({
                'Min': '{:.2f}',
                'Max': '{:.2f}'
            }),
            use_container_width=True
        )
    
    with tab4:
        top10_display = results['plots']['top10_data'][['Timestamp', 'Consumption [kWh]']].copy()
        st.dataframe(
            top10_display.style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True
        )
    
    with tab5:
        st.dataframe(
            results['plots']['ldc_data'].style.format({
                'Consumption [kWh]': '{:.2f}',
                'Percentile': '{:.2f}'
            }),
            use_container_width=True,
            height=400
        )

def create_export_options(results: Dict[str, Any]):
    """Provide multiple export formats"""
    
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        filename = f"energy_analysis_{timestamp_str}.xlsx"
        st.download_button(
            label="📥 Download Excel Report",
            data=results['excel_file'],
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col2:
        csv_buffer = io.StringIO()
        results['hourly_df'].to_csv(csv_buffer, index=False)
        st.download_button(
            label="📄 Download CSV (Hourly)",
            data=csv_buffer.getvalue(),
            file_name=f"hourly_data_{timestamp_str}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col3:
        json_data = results['hourly_df'].to_json(orient='records', date_format='iso')
        st.download_button(
            label="📋 Download JSON",
            data=json_data,
            file_name=f"data_{timestamp_str}.json",
            mime="application/json",
            use_container_width=True
        )
    
    st.info("💡 The Excel file includes all visualizations: heatmaps, monthly stats, seasonal patterns, and detailed analysis across multiple sheets.")


# ============================================
# MAIN STREAMLIT APPLICATION
# ============================================

def main():
    """Main Streamlit application"""
    
    initialize_session_state()
    
    # Header
    st.markdown("""
        <div class='main-title'>
            <h1>⚡ Energy Consumption Analyzer</h1>
        </div>
        <div class='subtitle'>
            Automated analysis of quarter-hourly energy data
        </div>
    """, unsafe_allow_html=True)
    
    # ============================================
    # SIDEBAR CONFIGURATION
    # ============================================
    
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # How to Use Section
        with st.expander("ℹ️ How to Use", expanded=False):
            st.markdown("""
            1. **Upload** your data file (Excel or CSV)
            2. **Select** date range for analysis
            3. **Click** "Start Analysis"
            4. **Download** comprehensive Excel report
            """)
        
        # Features Section
        with st.expander("📋 Features", expanded=False):
            st.markdown("""
            ✅ Data quality validation  
            ✅ Multiple visualization types  
            ✅ Hourly heatmaps (yearly, Jan, Jul)  
            ✅ Hourly and monthly summaries  
            ✅ Peak demand analysis  
            ✅ Load duration curves  
            ✅ Monthly min/max statistics  
            ✅ Comprehensive Excel report  
            ✅ Handles US and EU formats  
            ✅ Auto-detects file encoding
            """)
    
    # ============================================
    # MAIN CONTENT AREA
    # ============================================
    
    # Information Section - Two columns
    col1, col2 = st.columns(2)
    
    # LEFT COLUMN: Expected Data Format
    with col1:
        st.markdown("### ℹ️ Expected Data Format")
        st.markdown(f"""
        **Column structure:**
        - **Column 0**: Date
        - **Column 1**: Time  
        - **Column 2**: Consumption value
        
        **Supported formats:**
        - Date: DD.MM.YYYY, MM/DD/YYYY, YYYY-MM-DD
        - Time: HH:MM:SS, HH:MM
        - Numbers: Both US (.) and EU (,) decimal formats
        
        **File requirements:**
        - Maximum size: {Config.MAX_FILE_SIZE_MB}MB
        - Formats: {', '.join(Config.SUPPORTED_FORMATS)}
        """)
    
    # RIGHT COLUMN: Date Range Selection
    with col2:
        st.markdown("### 📅 Date Range Selection")
        
        date_selection = st.radio(
            "Select date range:",
            options=['all', 'recent', 'custom'],
            format_func=lambda x: {
                'all': '📊 Analyze all available data',
                'recent': '📅 Most recent full year only',
                'custom': '🎯 Specific year'
            }[x],
            index=0,
            label_visibility="collapsed"
        )
        
        selected_year = None
        if date_selection == 'custom':
            current_year = datetime.now().year
            selected_year = st.selectbox(
                "Select year:",
                options=list(range(current_year, 2019, -1)),
                index=0
            )
        else:
            selected_year = datetime.now().year
        
        # Display selected range info
        if date_selection == 'all':
            st.info("**Range:** All available data")
        elif date_selection == 'recent':
            st.info(f"**Range:** Most recent full year")
        else:
            st.info(f"**Range:** Year {selected_year}")
    
    st.markdown("---")
    
    # ============================================
    # UPLOAD SECTION
    # ============================================
    
    st.markdown("### 📁 Upload Files")
    
    st.markdown("**Choose your data file to process**")
    st.caption(f"Limit {Config.MAX_FILE_SIZE_MB}MB per file • {', '.join(Config.SUPPORTED_FORMATS).upper()}")
    
    uploaded_file = st.file_uploader(
        "Upload your data file",
        type=Config.SUPPORTED_FORMATS,
        label_visibility="collapsed"
    )
    
    if uploaded_file is not None:
        # File uploaded indicator
        st.markdown(f"""
            <div style='background: #d4edda; padding: 0.75rem; border-radius: 0.5rem; 
                        text-align: center; margin: 1rem 0; border-left: 4px solid #28a745;'>
                <strong style='color: #155724;'>✅ File uploaded: {uploaded_file.name}</strong>
            </div>
        """, unsafe_allow_html=True)
        
        file_valid = validate_uploaded_file(uploaded_file)
        
        st.markdown("---")
        
        # Analyze button
        analyze_button = st.button(
            "🚀 Start Analysis",
            type="primary",
            disabled=not file_valid,
            use_container_width=True
        )
    else:
        analyze_button = False
        file_valid = False
    
    # ============================================
    # PROCESSING AND RESULTS
    # ============================================
    
    if analyze_button and file_valid:
        file_content = uploaded_file.getvalue()
        current_hash = get_file_hash(file_content)
        
        if current_hash == st.session_state.last_file_hash and st.session_state.results is not None:
            st.info("ℹ️ Using cached results from previous analysis")
            results = st.session_state.results
            error = None
        else:
            with st.spinner("🔄 Processing your data..."):
                results, _, error = process_data_streamlit(
                    current_hash,
                    file_content,
                    uploaded_file.name,
                    date_selection, 
                    selected_year
                )
            
            if results is not None:
                st.session_state.results = results
                st.session_state.last_file_hash = current_hash
                st.session_state.file_processed = True
        
        if error:
            st.error(f"❌ {error}")
            st.stop()
        
        if results is None:
            st.error("❌ Failed to process data. Please check your file format.")
            st.stop()
        
        # Data Quality Report
        with st.expander("🔍 Data Quality Report", expanded=bool(results['issues'])):
            show_data_quality_report(results['issues'])
        
        st.success(f"✅ Analysis complete! {results['date_range_info']}")
        
        # Key Metrics
        st.markdown("### 📊 Key Metrics")
        display_summary_metrics(results)
        
        st.divider()
        
        # Results Tabs
        tab1, tab2, tab3 = st.tabs([
            "📊 Visualizations",
            "📋 Data Tables",
            "⬇️ Download"
        ])
        
        with tab1:
            display_visualizations(results['plots'])
        
        with tab2:
            display_data_tables(results)
        
        with tab3:
            st.markdown("### 📥 Download Reports")
            st.markdown("Download your analysis in multiple formats:")
            create_export_options(results)


if __name__ == "__main__":
    main()

